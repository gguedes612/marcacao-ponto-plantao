from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill,  Font

class Planilha:

    def __init__(self,diretorio):
        
        self.diretorio = diretorio
        self.font = Font('Calibri',color='FFFFFF')
        self.fill = PatternFill(patternType='solid',fgColor='4472c4')

        try: 
        
            self.workbook = load_workbook(self.diretorio)
            self.planilha = self.workbook['Planilha']
        
        except FileNotFoundError:
            
            self.workbook = Workbook()
            self.workbook.create_sheet('Planilha',0)
            self.planilha = self.workbook['Planilha']
            
            #Adicionando valores na planilha
            self.planilha['A1'] = 'Empresa'
            self.planilha['B1'] = 'Matricula'
            self.planilha['C1'] = 'Nome Completo'
            self.planilha['D1'] = 'Localidade'
            self.planilha['E1'] = 'Tipo'
            self.planilha['F1'] = 'Data Início'
            self.planilha['G1'] = 'D. Sem'
            self.planilha['H1'] = 'Hora Início'
            self.planilha['I1'] = 'Data Fim'
            self.planilha['J1'] = 'D. Sem'
            self.planilha['K1'] = 'Hora Fim'
            self.planilha['L1'] = 'T. Realizado (Hrs)'
            self.planilha['M1'] = 'T. Realizado (Num)'
            
            #Adicionando Fonte e Background
            self.planilha.row_dimensions[1].font = self.font
            self.planilha.row_dimensions[1].fill = self.fill


       
    def mostra_planilhas(self):
        return self.workbook.sheetnames

    def mostrar_celula(self,celula):
        return self.planilha[celula].value

    def salvar_planilha(self):
        self.workbook.save(self.diretorio)
    
    def adicionar_valores(self,nome_empresa,numero_matricula,nome_completo,localidade,data_inicio,hora_inicio,data_final,hora_final,linha=1):
        if self.planilha[f'A{linha}'].value == None:
            self.planilha[f'A{linha}'] = nome_empresa
            self.planilha[f'B{linha}'] = numero_matricula
            self.planilha[f'C{linha}'] = nome_completo
            self.planilha[f'D{linha}'] = localidade
            self.planilha[f'E{linha}'] = 'Acionamento'
            self.planilha[f'F{linha}'] = data_inicio
            self.planilha[f'G{linha}'] = f'=IF(F{linha}="","",WEEKDAY(F{linha}))'
            self.planilha[f'H{linha}'] = hora_inicio
            self.planilha[f'I{linha}'] = data_final
            self.planilha[f'J{linha}'] = f'=IF(I{linha}="","",WEEKDAY(I{linha}))'
            self.planilha[f'K{linha}'] = hora_final
            self.planilha[f'L{linha}'] = f'=IF(H{linha}="",0,(TEXT(I{linha},"dd/mm/aaaa")&" "&TEXT(K{linha},"[hh]:mm"))-(TEXT(F{linha},"dd/mm/aaaa")&" "&TEXT(H{linha},"[hh]:mm")))'
            self.planilha[f'M{linha}'] = f'=IF(L{linha}="",0,L{linha}*24)'
            
            self.adicionar_formato_celulas(linha)
            
        else:
            self.adicionar_valores(nome_empresa,numero_matricula,nome_completo,localidade,data_inicio,hora_inicio,data_final,hora_final,linha+1)

    def adicionar_formato_celulas(self,linha):
        self.planilha[f'G{linha}'].number_format = '[$-16]DDD'
        self.planilha[f'J{linha}'].number_format = '[$-16]DDD'
        self.planilha[f'L{linha}'].number_format = '[hh]:mm'
        self.planilha[f'M{linha}'].number_format = '0.00'
